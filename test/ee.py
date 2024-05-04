import openpyxl
from bot.other_methods.check_wagon_number import check_wagon_number

# Открываем файл Excel
# wb = openpyxl.load_workbook('/Users/szely/Downloads/2.xlsx')
# sheet = wb.active
#
# # Создаем новый файл Excel для записи результатов
# wb_result = openpyxl.Workbook()
# sheet_result = wb_result.active
#
# # Производим вычисления для каждой ячейки
# for row in sheet.iter_rows(values_only=True):
#     for cell in row:
#         # Пример вычисления: добавляем 10 к значению ячейки
#         result = check_wagon_number(cell)
#
#         # Записываем результат в следующую колонку
#         print(result)
#         sheet_result.cell(row=cell.row, column=cell.column + 1, value=result)
#
# # Сохраняем результаты в новый файл Excel
# wb_result.save('результат.xlsx')

import openpyxl

# Открываем Excel файл
wb = openpyxl.load_workbook('/Users/szely/Downloads/2.xlsx')
sheet = wb.active

# Проходим по каждой ячейке в первом столбце
for i in range(1, sheet.max_row + 1):
    cell_value = sheet.cell(row=i, column=1).value

    # Выполняем вычисления (в данном примере просто умножаем значение на 2)
    result = check_wagon_number(cell_value)

    # Записываем результат во второй столбец
    sheet.cell(row=i, column=2, value=result)

# Сохраняем изменения в файл
wb.save('output.xlsx')